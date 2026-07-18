"""Generate the small COREP LCR sample fact files (see demo/README.md).

The (template, row, column) combinations are real cells from the EBA DPM 2.0
v4.2 release (module COREP_LCR_DA), so the data resolves. All entity data is
fictional. Re-run to regenerate:

    python demo/build_demo.py

By default a run derives its filing indicators + parameters in-system, so only a
fact file is needed. ``indicators_params.xlsx`` is the optional "advanced"
override, paired with the warnings set below.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook

HERE = Path(__file__).parent

REFERENCE_DATE = date(2025, 12, 31)

# --- Minimal clean set — validates fully clean (derived indicators/params) ----
# C_76.00.a r0030 is a percentage datapoint; a ratio (<= 1) keeps it clean.
GOLDEN_ROWS = [
    ("C_73_00_a", "0010", "0010", 4500000),  # upstream underscore form
    ("C_73_00_a", "0020", "0010", 1250000),
    ("C_73_00_a", "0030", "0010", 830000),
    ("C 74.00.a", "0010", "0010", 2100000),  # EBA display (space) form
    ("C 74.00.a", "0020", "0010", 640000),
    ("C_76.00.a", "0010", "0010", 6800000),  # DB form
    ("C_76.00.a", "0020", "0010", 5400000),
    ("C_76.00.a", "0030", "0010", 0.87),  # ratio -> clean
]

# --- Warnings set — PERCENTAGE_NOT_RATIO + EMPTY_FILING_INDICATOR --------------
# Run WITH indicators_params.xlsx (override) to also surface the empty indicator.
WARNINGS_ROWS = [
    ("C_73.00.a", "0010", "0010", 4500000),
    ("C_73.00.a", "0020", "0010", 1250000),
    ("C_76.00.a", "0010", "0010", 6800000),
    ("C_76.00.a", "0030", "0010", 1.45),  # percentage > 1 -> PERCENTAGE_NOT_RATIO
]

# --- Malformed set — errors -> failed_validation ------------------------------
BROKEN_ROWS = [
    ("C_73.00.a", "0010", "0010", 4500000),  # ok
    ("C_73.00.a", "9999", "9999", 1),  # UNRESOLVED_FACT
    ("C_74.00.a", "0010", "0010", "not-a-number"),  # DATATYPE_MISMATCH (monetary)
]

# Advanced override: lists C_72.00.a (never has facts -> empty) plus the reported
# templates, so the warnings set shows EMPTY_FILING_INDICATOR.
OVERRIDE_INDICATORS = ["C_72.00.a", "C_73.00.a", "C_74.00.a", "C_76.00.a"]


def _fact_file(path: Path, rows: list) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "facts"
    ws.append(["report", "row", "column", "value"])
    for r in rows:
        ws.append(list(r))
    wb.save(path)


def build_indicators_params_file(path: Path) -> None:
    wb = Workbook()
    params = wb.active
    params.title = "parameters"
    params.append(["entity_lei", "213800MERIDNGRPHLD42"])  # fictional (Meridian)
    params.append(["reference_date", REFERENCE_DATE])
    params.append(["base_currency", "EUR"])
    params.append(["decimals", -3])

    indicators = wb.create_sheet("filing_indicators")
    indicators.append(["template", "reported"])
    for code in OVERRIDE_INDICATORS:
        indicators.append([code, True])
    wb.save(path)


def main() -> None:
    _fact_file(HERE / "fact_sample.xlsx", GOLDEN_ROWS)
    _fact_file(HERE / "fact_sample_warnings.xlsx", WARNINGS_ROWS)
    _fact_file(HERE / "fact_broken.xlsx", BROKEN_ROWS)
    build_indicators_params_file(HERE / "indicators_params.xlsx")
    for name in (
        "fact_sample.xlsx",
        "fact_sample_warnings.xlsx",
        "fact_broken.xlsx",
        "indicators_params.xlsx",
    ):
        print(f"wrote {HERE / name}")


if __name__ == "__main__":
    main()
