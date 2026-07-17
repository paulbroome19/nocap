"""Helpers to build in-memory XLSX bytes for tests (not collected as tests)."""

from __future__ import annotations

from collections.abc import Iterable, Sequence
from io import BytesIO

from openpyxl import Workbook


def fact_xlsx(
    rows: Iterable[Sequence[object]],
    *,
    header: Sequence[object] | None = ("report", "row", "column", "value"),
    sheet_title: str = "facts",
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    if header is not None:
        ws.append(list(header))
    for row in rows:
        ws.append(list(row))
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def indicators_params_xlsx(
    params: Iterable[Sequence[object]],
    indicators: Iterable[Sequence[object]],
    *,
    params_sheet: str = "parameters",
    indicators_sheet: str | None = "filing_indicators",
) -> bytes:
    wb = Workbook()
    p = wb.active
    p.title = params_sheet
    for row in params:
        p.append(list(row))
    if indicators_sheet is not None:
        fi = wb.create_sheet(indicators_sheet)
        fi.append(["template", "reported"])
        for row in indicators:
            fi.append(list(row))
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
