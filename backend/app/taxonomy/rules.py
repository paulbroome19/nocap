"""EBA validation-rules workbook: the ``validation_rules`` release slot.

The third *functional* artifact of a release (alongside the DPM database and the
taxonomy package). The EBA publishes its validation rules as a single-sheet
``.xlsx``; on upload we verify the header, retain the original byte-for-byte
(the seal), and ingest the rows into ``validation_rule`` as a background job —
exactly like the DPM Access→SQLite conversion. The slot's status lives on its
``ReleaseArtifact`` row (``verifying`` → ``ready`` / ``failed``).

Two things consume the ingested rules, both via ``workflows`` (never a direct
cross-stage import):

- the **register join** — a run's formula rows gain the workbook ``Description``,
  and deactivation is driven by ``IsActive`` + ``[From, To]ReferenceDate``
  evaluated against the run's reporting date (see ``build_register_view``);
- the **coherence check** — module-version tokens are cross-checked against the
  DPM / taxonomy-package versions.

Per the dependency rules this imports only ``app.core`` and its own stage.
"""

from __future__ import annotations

import io
import logging
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app.core.config import Settings, get_settings
from app.core.db import SessionLocal
from app.core.errors import ValidationError
from app.taxonomy.models import (
    ArtifactStatus,
    ReleaseArtifact,
    ReleaseSlot,
    TaxonomySnapshot,
    ValidationRule,
)
from app.taxonomy.service import compute_checksum, snapshot_dir

logger = logging.getLogger(__name__)

# The workbook's leading columns, in order. Extra trailing columns (Severity,
# FromSubmissionDate, Implemented in XBRL, …) are tolerated. The published file
# misspells "Precondition" as "Precondtion"; both spellings are accepted.
WORKBOOK_COLUMNS: tuple[str, ...] = (
    "VR Code",
    "Source",
    "Frameworks",
    "Modules",
    "Cross module",
    "Tables",
    "Expression",
    "Precondition",
    "Description",
    "IsActive",
    "FromReferenceDate",
    "ToReferenceDate",
)
_PRECONDITION_ALIASES = {"precondition", "precondtion"}


def _norm(value: object) -> str:
    return str(value or "").strip().casefold()


def _column_index(header: tuple, name: str) -> int | None:
    """Index of a column in the actual header row, by (case-insensitive) name.

    Robust to reordering and extra columns; resolves the Precondition spelling.
    """
    wanted = {name.casefold()}
    if name == "Precondition":
        wanted = _PRECONDITION_ALIASES
    for i, cell in enumerate(header):
        if _norm(cell) in wanted:
            return i
    return None


def verify_workbook_header(data: bytes) -> tuple:
    """Read + validate the workbook header row. Returns it, else raises.

    Verifies every expected leading column is present in order (tolerating the
    Precondition misspelling and any extra trailing columns). Raised errors are
    surfaced to the uploader verbatim, so they must be precise.
    """
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - openpyxl is a hard dep
        raise ValidationError(f"cannot read workbook: {exc}") from exc
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(data), read_only=True, data_only=True
        )
    except Exception as exc:  # noqa: BLE001 — any parse failure is a bad file
        raise ValidationError(f"not a readable .xlsx workbook: {exc}") from exc
    try:
        ws = wb[wb.sheetnames[0]]
        header = next(ws.iter_rows(values_only=True), None)
    finally:
        wb.close()
    if not header:
        raise ValidationError("workbook is empty (no header row)")
    for i, expected in enumerate(WORKBOOK_COLUMNS):
        aliases = (
            _PRECONDITION_ALIASES
            if expected == "Precondition"
            else {expected.casefold()}
        )
        if i >= len(header) or _norm(header[i]) not in aliases:
            got = _norm(header[i]) if i < len(header) else "(missing)"
            raise ValidationError(
                f"unexpected validation-rules workbook: column {i + 1} should be "
                f"{expected!r}, got {got!r}. This does not look like the EBA "
                "validation-rules file."
            )
    return header


def verify_workbook_file(data: bytes, filename: str) -> None:
    """Verify an uploaded validation-rules workbook, or raise plainly.

    Extension + header check, with EBA-website-term messaging so a reporter
    knows which download to grab.
    """
    from pathlib import Path

    if not data:
        raise ValidationError("The validation-rules workbook is empty.")
    if Path(filename).suffix.lower() != ".xlsx":
        raise ValidationError(
            "This is not the EBA validation-rules workbook. On the EBA "
            "reporting-frameworks page, download the validation rules for this "
            "release — an Excel .xlsx — and upload that."
        )
    try:
        verify_workbook_header(data)  # raises on unreadable / wrong columns
    except ValidationError as exc:
        raise ValidationError(
            f"This is not the EBA validation-rules workbook — {exc}"
        ) from exc


# ---------------------------------------------------------------------------
# Storage + ingestion (background, status-tracked like the DPM conversion)
# ---------------------------------------------------------------------------


def _rules_dir(settings: Settings, snapshot_id: int) -> Path:
    return snapshot_dir(settings, snapshot_id) / "rules"


def store_workbook(
    db: Session,
    snapshot: TaxonomySnapshot,
    *,
    filename: str,
    data: bytes,
    settings: Settings | None = None,
) -> ReleaseArtifact:
    """Accept + seal a validation-rules workbook, marking its slot ``verifying``.

    Verifies the header synchronously (so a wrong file is rejected at upload),
    writes the original byte-for-byte, and upserts the slot's ``ReleaseArtifact``
    to ``verifying``. The caller schedules ``ingest_validation_rules_task`` to
    parse the rows in the background.
    """
    settings = settings or get_settings()
    if not data:
        raise ValidationError("uploaded file is empty")
    if Path(filename).suffix.lower() != ".xlsx":
        raise ValidationError(
            f"validation rules expect .xlsx (got {Path(filename).suffix or 'no'} "
            "extension)"
        )
    verify_workbook_header(data)  # raises on a wrong/unreadable file

    target_dir = _rules_dir(settings, snapshot.id)
    if target_dir.exists():
        for old in target_dir.glob("*.xlsx"):
            old.unlink()
    target_dir.mkdir(parents=True, exist_ok=True)
    path = target_dir / filename
    path.write_bytes(data)

    artifact = db.scalar(
        select(ReleaseArtifact).where(
            ReleaseArtifact.snapshot_id == snapshot.id,
            ReleaseArtifact.slot == ReleaseSlot.validation_rules,
        )
    )
    if artifact is None:
        artifact = ReleaseArtifact(
            snapshot_id=snapshot.id, slot=ReleaseSlot.validation_rules
        )
        db.add(artifact)
    artifact.filename = filename
    artifact.storage_key = str(path.relative_to(settings.data_dir))
    artifact.checksum = compute_checksum(data)
    artifact.status = ArtifactStatus.verifying
    artifact.error = None
    db.commit()
    db.refresh(artifact)
    logger.info(
        "stored validation_rules workbook for release id=%s (verifying)",
        snapshot.id,
    )
    return artifact


def _as_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    # openpyxl yields datetimes for date cells; strings like "NULL"/"" are empty.
    return None


def _as_bool_active(value: object) -> bool:
    # "Active" / "Inactive" (case-insensitive); anything else defaults to active.
    return not _norm(value).startswith("i")


def parse_workbook_rows(snapshot_id: int, data: bytes) -> list[dict]:
    """Parse the workbook into ``ValidationRule`` insert mappings (pure)."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None) or ()
        idx = {
            key: _column_index(header, col)
            for key, col in (
                ("vr_code", "VR Code"),
                ("source", "Source"),
                ("frameworks", "Frameworks"),
                ("modules", "Modules"),
                ("cross_module", "Cross module"),
                ("tables", "Tables"),
                ("expression", "Expression"),
                ("precondition", "Precondition"),
                ("description", "Description"),
                ("is_active", "IsActive"),
                ("from_reference_date", "FromReferenceDate"),
                ("to_reference_date", "ToReferenceDate"),
                ("severity", "Severity"),
            )
        }

        def cell(row: tuple, key: str) -> object:
            i = idx.get(key)
            return row[i] if i is not None and i < len(row) else None

        def text(row: tuple, key: str) -> str | None:
            v = cell(row, key)
            if v is None:
                return None
            s = str(v).strip()
            return s or None

        mappings: list[dict] = []
        for row in rows:
            code = text(row, "vr_code")
            if not code:
                continue  # skip blank/trailing rows
            mappings.append(
                {
                    "snapshot_id": snapshot_id,
                    "vr_code": code,
                    "source": text(row, "source"),
                    "frameworks": text(row, "frameworks"),
                    "modules": text(row, "modules"),
                    "cross_module": text(row, "cross_module"),
                    "tables": text(row, "tables"),
                    "expression": text(row, "expression"),
                    "precondition": text(row, "precondition"),
                    "description": text(row, "description"),
                    "is_active": _as_bool_active(cell(row, "is_active")),
                    "from_reference_date": _as_date(
                        cell(row, "from_reference_date")
                    ),
                    "to_reference_date": _as_date(cell(row, "to_reference_date")),
                    "severity": text(row, "severity"),
                }
            )
        return mappings
    finally:
        wb.close()


def ingest_validation_rules(
    db: Session,
    snapshot: TaxonomySnapshot,
    *,
    settings: Settings | None = None,
) -> None:
    """Parse the sealed workbook into ``validation_rule`` rows; set slot status.

    Replaces any previously-ingested rules for the release (re-ingest is safe).
    Never raises — records failure on the slot's ``ReleaseArtifact`` row.
    """
    settings = settings or get_settings()
    artifact = db.scalar(
        select(ReleaseArtifact).where(
            ReleaseArtifact.snapshot_id == snapshot.id,
            ReleaseArtifact.slot == ReleaseSlot.validation_rules,
        )
    )
    if artifact is None:
        logger.warning(
            "validation_rules ingest: no slot row for release id=%s", snapshot.id
        )
        return
    try:
        path = settings.data_dir / artifact.storage_key
        if not path.exists():
            raise ValidationError("stored workbook is missing on disk")
        mappings = parse_workbook_rows(snapshot.id, path.read_bytes())
        if not mappings:
            raise ValidationError("workbook contains no validation rules")
        db.execute(
            delete(ValidationRule).where(
                ValidationRule.snapshot_id == snapshot.id
            )
        )
        db.bulk_insert_mappings(ValidationRule, mappings)
        artifact.status = ArtifactStatus.ready
        artifact.error = None
        db.commit()
        logger.info(
            "release id=%s ingested %d validation rules",
            snapshot.id, len(mappings),
        )
    except Exception as exc:  # noqa: BLE001 — record any failure on the slot
        db.rollback()
        artifact.status = ArtifactStatus.failed
        artifact.error = str(exc)[:2000]
        db.commit()
        logger.exception(
            "validation_rules ingest failed for release id=%s", snapshot.id
        )


def ingest_validation_rules_task(snapshot_id: int) -> None:
    """Background entrypoint: open a fresh session and ingest by release id."""
    settings = get_settings()
    with SessionLocal() as db:
        snapshot = db.get(TaxonomySnapshot, snapshot_id)
        if snapshot is None:
            logger.warning(
                "validation_rules ingest task: release id=%s not found", snapshot_id
            )
            return
        ingest_validation_rules(db, snapshot, settings=settings)


# ---------------------------------------------------------------------------
# Read: register join + coherence source
# ---------------------------------------------------------------------------


def has_ingested_rules(db: Session, snapshot_id: int) -> bool:
    """True if this release has any ingested validation rules."""
    return (
        db.scalar(
            select(ValidationRule.id)
            .where(ValidationRule.snapshot_id == snapshot_id)
            .limit(1)
        )
        is not None
    )


def _covers(rule: ValidationRule, reporting_date: date) -> bool:
    """Whether the rule's ``[from, to]`` reference-date window covers the date.

    Bounds are inclusive; an open bound (NULL) is unbounded on that side.
    """
    frm, to = rule.from_reference_date, rule.to_reference_date
    if frm is not None and reporting_date < frm:
        return False
    if to is not None and reporting_date > to:
        return False
    return True


@dataclass(frozen=True)
class RegisterRuleView:
    """The workbook facts a run's formula register needs, resolved for a date.

    - ``descriptions`` — vr_code → human rule statement (the covering row's, else
      any row's) for joining onto executed formula rows.
    - ``inactive`` — vr_code → description for codes that are NOT active for the
      reporting date (inactive, or outside their window) — used to flag rules the
      taxonomy executed but the workbook has deactivated.
    - ``deactivated_codes`` — the full set to exclude from the Arelle results for
      this reporting date. Large (most codes are out-of-window); membership only.
    - ``severities`` — vr_code → EBA severity of the covering row (e.g.
      ``"error"`` / ``"warning"``), the authoritative source for whether a
      failing formula rule is blocking.
    """

    descriptions: dict[str, str]
    inactive: dict[str, str]
    deactivated_codes: set[str] = field(default_factory=set)
    severities: dict[str, str] = field(default_factory=dict)


def build_register_view(
    db: Session, snapshot_id: int, reporting_date: date
) -> RegisterRuleView:
    """Resolve the ingested rules for a run's reporting date."""
    rows = list(
        db.scalars(
            select(ValidationRule).where(
                ValidationRule.snapshot_id == snapshot_id
            )
        )
    )
    all_codes: set[str] = set()
    any_desc: dict[str, str] = {}
    # Effective covering row per code (prefer active, then the latest window).
    best: dict[str, ValidationRule] = {}

    def rank(r: ValidationRule) -> tuple:
        return (r.is_active, r.from_reference_date or date.min)

    for r in rows:
        all_codes.add(r.vr_code)
        if r.description and r.vr_code not in any_desc:
            any_desc[r.vr_code] = r.description
        if _covers(r, reporting_date):
            cur = best.get(r.vr_code)
            if cur is None or rank(r) > rank(cur):
                best[r.vr_code] = r

    active_codes = {c for c, r in best.items() if r.is_active}
    descriptions = dict(any_desc)
    severities: dict[str, str] = {}
    for c, r in best.items():
        if r.description:
            descriptions[c] = r.description  # covering row's wins
        if r.severity:
            severities[c] = r.severity.strip().lower()

    inactive_codes = all_codes - active_codes
    inactive = {c: any_desc.get(c, "") for c in inactive_codes}
    return RegisterRuleView(
        descriptions=descriptions,
        inactive=inactive,
        deactivated_codes=inactive_codes,
        severities=severities,
    )
