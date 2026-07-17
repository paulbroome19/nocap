"""Defensive parsers for the two input files (see CLAUDE.md "Input contracts").

Shape-only: these confirm the files are parseable, non-empty, and that codes are
well-formed, producing precise row-level errors. Datapoint resolution and
datatype checks belong to the validation stage — this keeps that seam clean.

The template-code normaliser is *injected* (``TemplateNormalizer``): the facts
stage must not import the taxonomy stage. ``workflows`` (and the app composition
root) supplies the taxonomy contract's normaliser.
"""

from __future__ import annotations

import datetime as dt
from collections.abc import Callable
from io import BytesIO
from typing import Protocol

from openpyxl import load_workbook

from app.facts.schemas import (
    FactFileParseResult,
    FilingIndicator,
    IndicatorsParams,
    IndicatorsParamsParseResult,
    ParsedFact,
    RowError,
)

# code (any input form) -> canonical code; raises ValueError if unrecognisable.
TemplateNormalizer = Callable[[str], str]

_FACT_COLUMNS = {
    "report": {"report", "template", "report code", "template code"},
    "row": {"row", "row code"},
    "column": {"column", "col", "column code"},
    "value": {"value", "amount"},
}


def _scalar_to_str(value: object) -> str:
    """Render a cell value as a faithful string (no leading-zero loss handled here)."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else repr(value)
    if isinstance(value, (dt.datetime, dt.date)):
        return value.isoformat()
    return str(value).strip()


def _at(cells: list[object], index: int) -> object:
    return cells[index] if index < len(cells) else None


def _code_str(value: object) -> str:
    """Render a row/column code as text, re-padding numeric-origin codes.

    Excel silently turns ``0010`` into the number ``10``; EBA row/column codes
    are 4-digit, so an all-digit code shorter than 4 is left-padded with zeros.
    """
    text = _scalar_to_str(value)
    if text.isdigit() and len(text) < 4:
        return text.zfill(4)
    return text


def _header_index(header_cells: list[object]) -> dict[str, int] | None:
    """Map the fact columns to their 0-based indices from a header row."""
    normalized = {
        i: _scalar_to_str(c).lower() for i, c in enumerate(header_cells)
    }
    index: dict[str, int] = {}
    for key, aliases in _FACT_COLUMNS.items():
        for i, name in normalized.items():
            if name in aliases:
                index[key] = i
                break
    return index if len(index) == len(_FACT_COLUMNS) else None


def parse_fact_xlsx(
    data: bytes, *, normalize: TemplateNormalizer
) -> FactFileParseResult:
    """Parse a fact XLSX (columns: report, row, column, value)."""
    facts: list[ParsedFact] = []
    errors: list[RowError] = []

    try:
        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — surface any unreadable file cleanly
        return FactFileParseResult(
            facts=[],
            errors=[RowError(sheet="?", row=0, message=f"unreadable XLSX: {exc}")],
        )

    ws = wb.worksheets[0]
    sheet = ws.title
    rows = ws.iter_rows(values_only=True)

    # Locate the header row (first row with any content).
    header_index: dict[str, int] | None = None
    header_row_no = 0
    for row_no, cells in enumerate(rows, start=1):
        if all(c is None or _scalar_to_str(c) == "" for c in cells):
            continue
        header_index = _header_index(list(cells))
        header_row_no = row_no
        break

    if header_index is None:
        errors.append(
            RowError(
                sheet=sheet,
                row=header_row_no or 1,
                message="missing columns: expected report, row, column, value",
            )
        )
        wb.close()
        return FactFileParseResult(facts=facts, errors=errors)

    for row_no, cells in enumerate(rows, start=header_row_no + 1):
        cell_list = list(cells)
        report = _scalar_to_str(_at(cell_list, header_index["report"]))
        value = _scalar_to_str(_at(cell_list, header_index["value"]))
        row_code = _code_str(_at(cell_list, header_index["row"]))
        column_code = _code_str(_at(cell_list, header_index["column"]))

        if not any([report, row_code, column_code, value]):
            continue  # fully blank row

        missing = [
            name
            for name, present in (
                ("report", report),
                ("row", row_code),
                ("column", column_code),
                ("value", value),
            )
            if not present
        ]
        if missing:
            errors.append(
                RowError(
                    sheet=sheet,
                    row=row_no,
                    message=f"empty required cell(s): {', '.join(missing)}",
                )
            )
            continue

        try:
            template_code = normalize(report)
        except ValueError:
            errors.append(
                RowError(
                    sheet=sheet,
                    row=row_no,
                    column="report",
                    message=f"unrecognised template code {report!r}",
                )
            )
            continue

        facts.append(
            ParsedFact(
                template_code=template_code,
                template_code_raw=report,
                row_code=row_code,
                column_code=column_code,
                value=value,
                source_sheet=sheet,
                source_row=row_no,
            )
        )

    wb.close()
    return FactFileParseResult(facts=facts, errors=errors)


# --- Indicators & parameters (behind an interface) -------------------------


class IndicatorsParamsParser(Protocol):
    """Interface so the exact layout can change without touching generation."""

    def parse(
        self, data: bytes, *, normalize: TemplateNormalizer
    ) -> IndicatorsParamsParseResult: ...


_REQUIRED_PARAM_KEYS = ("entity_lei", "reference_date", "base_currency", "decimals")


def _coerce_date(value: object) -> dt.date:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    return dt.date.fromisoformat(_scalar_to_str(value))


class XlsxIndicatorsParamsParser:
    """Default layout: a ``parameters`` sheet (key/value) and a
    ``filing_indicators`` sheet (``template`` [, ``reported``]).
    """

    def parse(
        self, data: bytes, *, normalize: TemplateNormalizer
    ) -> IndicatorsParamsParseResult:
        errors: list[RowError] = []
        try:
            wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001
            return IndicatorsParamsParseResult(
                params=None,
                errors=[RowError(sheet="?", row=0, message=f"unreadable XLSX: {exc}")],
            )

        names = {n.lower(): n for n in wb.sheetnames}
        for required in ("parameters", "filing_indicators"):
            if required not in names:
                errors.append(
                    RowError(
                        sheet=required,
                        row=0,
                        message=f"missing required sheet {required!r}",
                    )
                )
        if errors:
            wb.close()
            return IndicatorsParamsParseResult(params=None, errors=errors)

        params_dict, params_errors = self._read_parameters(wb[names["parameters"]])
        errors.extend(params_errors)
        indicators, ind_errors = self._read_indicators(
            wb[names["filing_indicators"]], normalize
        )
        errors.extend(ind_errors)
        wb.close()

        if errors:
            return IndicatorsParamsParseResult(params=None, errors=errors)

        return IndicatorsParamsParseResult(
            params=IndicatorsParams(
                filing_indicators=indicators,
                entity_lei=params_dict["entity_lei"],
                reference_date=params_dict["reference_date"],
                base_currency=params_dict["base_currency"],
                decimals=params_dict["decimals"],
            ),
            errors=[],
        )

    def _read_parameters(self, ws) -> tuple[dict, list[RowError]]:
        sheet = ws.title
        errors: list[RowError] = []
        # Keep the raw cell value (not stringified) so typed cells — e.g. an
        # Excel date — reach the field coercers with their type intact.
        raw: dict[str, tuple[int, object]] = {}
        for row_no, cells in enumerate(ws.iter_rows(values_only=True), start=1):
            if not cells or _scalar_to_str(cells[0]) == "":
                continue
            key = _scalar_to_str(cells[0]).lower()
            value = cells[1] if len(cells) > 1 else None
            raw[key] = (row_no, value)

        out: dict[str, object] = {}
        for key in _REQUIRED_PARAM_KEYS:
            if key not in raw:
                errors.append(
                    RowError(sheet=sheet, row=0, message=f"missing parameter {key!r}")
                )
        if errors:
            return out, errors

        lei_row, lei_raw = raw["entity_lei"]
        lei = _scalar_to_str(lei_raw)
        if len(lei) != 20 or not lei.isalnum():
            errors.append(
                RowError(
                    sheet=sheet,
                    row=lei_row,
                    column="entity_lei",
                    message=f"malformed LEI {lei!r} (expected 20 alphanumeric chars)",
                )
            )
        else:
            out["entity_lei"] = lei.upper()

        date_row, date_val = raw["reference_date"]
        try:
            out["reference_date"] = _coerce_date(date_val)
        except (ValueError, TypeError):
            errors.append(
                RowError(
                    sheet=sheet,
                    row=date_row,
                    column="reference_date",
                    message=f"unparseable reference_date {date_val!r}",
                )
            )

        cur_row, cur_raw = raw["base_currency"]
        currency = _scalar_to_str(cur_raw)
        if len(currency) != 3 or not currency.isalpha():
            errors.append(
                RowError(
                    sheet=sheet,
                    row=cur_row,
                    column="base_currency",
                    message=(
                        f"malformed base_currency {currency!r} (expected 3 letters)"
                    ),
                )
            )
        else:
            out["base_currency"] = currency.upper()

        dec_row, dec_raw = raw["decimals"]
        try:
            out["decimals"] = int(float(_scalar_to_str(dec_raw)))
        except (ValueError, TypeError):
            errors.append(
                RowError(
                    sheet=sheet,
                    row=dec_row,
                    column="decimals",
                    message=f"non-integer decimals {dec_raw!r}",
                )
            )

        return out, errors

    def _read_indicators(
        self, ws, normalize: TemplateNormalizer
    ) -> tuple[list[FilingIndicator], list[RowError]]:
        sheet = ws.title
        errors: list[RowError] = []
        indicators: list[FilingIndicator] = []

        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        # Skip a header row if present (first cell not a template code).
        if header is not None and _scalar_to_str(header[0]).lower() in {
            "template",
            "template code",
            "report",
        }:
            start = 2
        else:
            rows = ws.iter_rows(values_only=True)  # rewind: no header
            start = 1

        for row_no, cells in enumerate(rows, start=start):
            if not cells or _scalar_to_str(cells[0]) == "":
                continue
            code = _scalar_to_str(cells[0])
            reported = True
            if len(cells) > 1 and _scalar_to_str(cells[1]) != "":
                reported = _scalar_to_str(cells[1]).lower() in {
                    "true",
                    "1",
                    "yes",
                    "y",
                }
            try:
                indicators.append(
                    FilingIndicator(template_code=normalize(code), reported=reported)
                )
            except ValueError:
                errors.append(
                    RowError(
                        sheet=sheet,
                        row=row_no,
                        column="template",
                        message=f"unrecognised template code {code!r}",
                    )
                )

        if not indicators and not errors:
            errors.append(
                RowError(sheet=sheet, row=0, message="no filing indicators found")
            )
        return indicators, errors


# Default parser instance used by the service (swappable per the interface).
default_indicators_params_parser: IndicatorsParamsParser = (
    XlsxIndicatorsParamsParser()
)
